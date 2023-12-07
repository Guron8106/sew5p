import logging
import random
import string
import argparse
import re
import unicodedata
from collections import namedtuple

from logging.handlers import RotatingFileHandler
from openpyxl import load_workbook


def shave_marks(txt):
    """Remove all diacritic marks"""
    norm_txt = unicodedata.normalize('NFD', txt)
    shaved = ''.join(c for c in norm_txt
                     if not unicodedata.combining(c))
    erg = unicodedata.normalize('NFC', shaved)
    return erg


def generate_password():
    """Generates Password for a user
    for _ in range(10): This part creates a loop that repeats 10 times (range(10)). During each iteration,
     a random letter is chosen.

    """
    password_chars = string.ascii_letters + "!%&(),._-=^#!%&(),._-=^#"
    logger.debug("generated password")
    return ''.join(random.choice(password_chars) for _ in range(10))


def get_user():
    """Generator for user
    namedTubles -> subclass
    The second argument is a string containing the names of the fields
    separated by spaces or commas ("vname nname group u_class login_name").
    """
    User = namedtuple("User", "vname nname group u_class login_name")
    for row in ws.iter_rows(min_row=2):
        firstname = shave_marks(str(row[0].value).lower())
        lastname = str(row[1].value).lower()
        # Replacing specific characters
        replacements = {'ö': 'oe', 'ä': 'ae', 'ü': 'ue', 'ß': 'ss'}
        for char, replacement in replacements.items():
            lastname = lastname.replace(char, replacement)
        lastname = shave_marks(lastname)
        group = str(row[2].value)
        u_class = str(row[3].value)
        user = User(firstname, lastname, group, u_class, "")
        yield user


def generate_scripts():
    """Generates scripts and iterates through the users

                    login_name = re.sub(r"(\d+)", "", login_name)
                -> darf nicht mit Zahl anfangen

            login_name = re.sub(r"(\d+)", "", login_name)
            login_name += str(counter)
            counter += 1
            -> Damit die Benutzernamen unique sind
            -> siehe Shell Skript
    """
    with open("res/create_user.sh", "w") as file:
        logger.debug("opened file " + file.name)
        print("set -e", file=file)
    with open("res/delete_user.sh", "w") as file:
        logger.debug("opened file " + file.name)
        print("set -x", file=file)
    open("res/passwords_user.txt", "w").close()
    users = dict()
    for user in get_user():
        login_name = user.nname
        counter = 1
        while login_name in users:
            login_name = re.sub(r"(\d+)", "", login_name)
            login_name += str(counter)
            counter += 1
        users[login_name] = login_name
        pw = generate_password()
        user = user._replace(login_name=login_name)
        create_user_entry(user, pw)


def create_user_entry(user, pw):
    """
    Creates a line in the create, delete script and in the password file for the given user

    """
    useradd(user, pw)
    userdel(user)
    addpasswd(user, pw)


def userdel(user):
    """Writes userdel command in respective File"""
    with open("res/delete_user.sh", "a") as file:
        logger.debug("opened file " + file.name)
        delete = f'userdel {user.login_name} && rm -rf /home/klassen/{user.login_name}'
        print(delete, file=file)
        logger.info("wrote userdel into " + file.name)


def useradd(user, pw):
    """Writes useradd command in respective File"""
    create = f'useradd -d "/home/{user.login_name}" -c "{user.vname + " " + user.nname}" -m ' \
             f'-g {user.group}{"," + user.u_class if user.group == "student" else ""} -s "/bin/bash {user.login_name}" && ' \
             f'echo {user.login_name}:\"{pw}\" | chpasswd'
    with open("res/create_user.sh", "a", encoding="utf-8") as file:
        logger.debug("opened file " + file.name)
        print(create, file=file)
        logger.info("wrote useradd into " + file.name)


def addpasswd(user, pw):
    """Writes user with their password in respective File"""
    with open("res/passwords_user.txt", "a") as file:
        logger.debug("opened file " + file.name)
        print(user.login_name, pw, file=file, sep=":")
        logger.info("wrote password into file for user in " + file.name)

if __name__ == '__main__':
    wb = load_workbook("Namen.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    logger = logging.getLogger()

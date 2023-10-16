import argparse
import logging
import random
import string
import sys
import unicodedata
from logging.handlers import RotatingFileHandler

from openpyxl.reader.excel import load_workbook

def shave_marks(txt: str):
    """
    Remove all diacritic marks
    """
    norm_txt = unicodedata.normalize('NFD', txt)
    shaved = ''.join(c for c in norm_txt
                     if not unicodedata.combining(c))
    return unicodedata.normalize('NFC', shaved)


def generate_password(user: tuple):
    """
    Generates Password for a user
    """
    special_chars = "!%&(),._-=^#!%&(),._-=^#"
    logger.debug("generated password")
    return f'{user[0]}{random.choice(special_chars)}{user[1]}' \
           f'{random.choice(special_chars)}{user[2]}{random.choice(special_chars)}'


def get_user():
    """
    Generator for user
    min_row = 2, damit die erste Row mit den Übrschriften übersprungen wird
    to use user as a tuple
    """
    for row in ws.iter_rows(min_row=2):
        klasse = str(row[0].value).lower()
        raum = str(row[1].value)
        kv = str(row[2].value)
        yield klasse, raum, kv


def generate_scripts():
    """
    Generates scripts and iterates through the users
    set -e = errexit
    set -x = xtrace debugging
    for _ in range(10): This part creates a loop that repeats 10 times (range(10)). During each iteration, a random letter is chosen.
    ''.join(...): This part joins the random letters generated in each iteration together into a single string. ''
    """
    with open("res/create_class.sh", "w") as file:
        logger.debug("opened file " + file.name)
        print("set -e", file=file)
    with open("res/delete_class.sh", "w") as file:
        logger.debug("opened file " + file.name)
        print("set -x", file=file)
    open("res/passwords_class.txt", "w").close()
    create_user_entry(("lehrer",), ''.join(random.choice(string.ascii_letters) for _ in range(10)))
    create_user_entry(("seminar",), ''.join(random.choice(string.ascii_letters) for _ in range(10)))
    for user in get_user():
        pw = generate_password(user)
        create_user_entry(user, pw)


def create_user_entry(user, pw):
    """
    Creates a line in the create, delete script and in the password file for the given user
    """
    useradd(user, pw)
    userdel(user)
    addpasswd(user, pw)


def userdel(user):
    """
    Writes userdel command in  File
    """
    delete = f'userdel {user[0]} && rm -rf /home/klassen/k{user[0]}'
    with open("res/delete_class.sh", "a") as file:
        logger.debug("opened file " + file.name)
        print(delete, file=file)
        logger.info("wrote userdel into " + file.name)


def useradd(user, pw):
    """
    Writes useradd command in respective File
    user[0][0] accesses the first character of the first element in the user tuple.
    user[0] = klasse
    """
    create = f'useradd -d /home/klassen/{"k" if user[0][0].isdigit() else ""}{user[0]} -c "{user[0]}" -m ' \
             f'-g cdrom,plugdev,sambashare -s /bin/bash {user[0]} && ' \
             f'echo {user[0]}:\"{pw}\" | chpasswd'
    with open("res/create_class.sh", "a") as file:
        logger.debug("opened file " + file.name)
        print(create, file=file)
        logger.info("wrote useradd into " + file.name)


def addpasswd(user, pw):
    """
    Writes user with their password in respective File
    """
    with open("res/passwords_class.txt", "a") as file:
        logger.debug("opened file " + file.name)
        print(user[0], pw, file=file, sep=":")
        logger.info("wrote password into file for user in " + file.name)



if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    output_group = parser.add_mutually_exclusive_group()
    output_group.add_argument("-v", "--verbose", help="increase output verbosity", action="store_true")
    output_group.add_argument("-q", "--quiet", help="decrease output verbosity", action="store_true")
    parser.add_argument("file", help="file with the userdata")
    args = parser.parse_args()

    logger = logging.getLogger()

    formatter = logging.Formatter("%(asctime)s; %(levelname)s; %(message)s",
                                  "%Y-%m-%d %H:%M:%S")

    rotating_file_handler = RotatingFileHandler("res/create_class.log", maxBytes=10000, backupCount=5)
    rotating_file_handler.setFormatter(formatter)
    logger.addHandler(rotating_file_handler)

    stream_handler = logging.StreamHandler(sys.stderr)
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    if args.verbose:
        stream_handler.setLevel(logging.DEBUG)
    elif args.quiet:
        stream_handler.setLevel(logging.CRITICAL)
    else:
        stream_handler.setLevel(logging.INFO)

    logger.setLevel(logging.DEBUG)

    try:
        # wb = load_workbook(args.file, read_only=True)
        wb = load_workbook("Klassenraeume_2023.xlsx", read_only=True)
        ws = wb[wb.sheetnames[0]]
        generate_scripts()
    except FileNotFoundError:
        logger.critical("couldnt find file")
    except Exception as e:
        logger.exception("Exception occurred: %s", e)

